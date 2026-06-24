"""Generate Excel workbook with newsletter data."""
import sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import json
import openpyxl
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint


def hex_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def create_excel(deals_path: str, output_path: str):
    with open(deals_path, encoding="utf-8") as f:
        deals = json.load(f)

    included = [d for d in deals if d.get("include_in_newsletter")]

    wb = openpyxl.Workbook()

    # ── Sheet 1: Newsletter Summary ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Newsletter Summary"

    # Title block
    ws1.merge_cells("A1:J1")
    ws1["A1"] = "🌍  FMCG M&A INTELLIGENCE NEWSLETTER  |  June 2026"
    ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws1["A1"].fill = hex_fill("1B3A5C")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:J2")
    ws1["A2"] = "Powered by FMCG Intel Agent  |  Real-time M&A Deal Tracker  |  For internal use only"
    ws1["A2"].font = Font(italic=True, size=10, color="888888")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # KPI Row
    total_val = sum(d.get("deal_value_usd_bn", 0) or 0 for d in included)
    kpis = [
        ("Deals Tracked", str(len(included))),
        ("Total Disclosed Value", f"${total_val:.1f}B"),
        ("Completed", str(sum(1 for d in included if "Completed" in str(d.get("status","")) or "Approved" in str(d.get("status",""))))),
        ("Pending/Announced", str(sum(1 for d in included if "Announced" in str(d.get("status",""))))),
        ("Failed", str(sum(1 for d in included if "Failed" in str(d.get("status","")) or "Shelved" in str(d.get("status",""))))),
    ]
    ws1.row_dimensions[3].height = 10
    kpi_colors = ["0D7377", "14A085", "32E0C4", "F0A500", "C0392B"]
    for i, (label, val) in enumerate(kpis):
        col = i * 2 + 1
        ws1.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        ws1.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        cell_label = ws1.cell(row=4, column=col, value=label)
        cell_val = ws1.cell(row=5, column=col, value=val)
        cell_label.font = Font(bold=True, size=9, color="FFFFFF")
        cell_val.font = Font(bold=True, size=18, color="FFFFFF")
        cell_label.fill = hex_fill(kpi_colors[i])
        cell_val.fill = hex_fill(kpi_colors[i])
        cell_label.alignment = Alignment(horizontal="center", vertical="center")
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[4].height = 20
        ws1.row_dimensions[5].height = 32

    ws1.row_dimensions[6].height = 10

    # Deal table header
    headers = ["#", "Headline", "Acquirer", "Target", "Value ($B)", "Category",
               "Geography", "Status", "Rel. Score", "Source"]
    header_row = 7
    for j, h in enumerate(headers, 1):
        c = ws1.cell(row=header_row, column=j, value=h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = hex_fill("1B3A5C")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws1.row_dimensions[header_row].height = 22

    status_colors = {
        "Completed": "D5F5E3",
        "Approved": "D5F5E3",
        "Announced": "FEF9E7",
        "Failed": "FADBD8",
        "Shelved": "FADBD8",
    }

    for idx, deal in enumerate(included, 1):
        row = header_row + idx
        val = deal.get("deal_value_usd_bn")
        val_str = f"{val:.1f}" if val else "N/D"
        status = deal.get("status", "")
        bg = "FFFFFF"
        for k, color in status_colors.items():
            if k in status:
                bg = color
                break

        row_data = [
            idx,
            deal.get("headline", ""),
            deal.get("acquirer", ""),
            deal.get("target", ""),
            val_str,
            deal.get("category", ""),
            deal.get("geography", ""),
            status,
            deal.get("composite_score", ""),
            deal.get("source", ""),
        ]
        for j, val_d in enumerate(row_data, 1):
            c = ws1.cell(row=row, column=j, value=val_d)
            c.fill = hex_fill(bg)
            c.font = Font(size=9)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border = thin_border()
        ws1.row_dimensions[row].height = 28

    # Column widths
    col_widths = [4, 42, 18, 20, 10, 22, 14, 18, 10, 22]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Pipeline Log ────────────────────────────────────────
    ws2 = wb.create_sheet("Pipeline Log")
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "Pipeline Transparency Log"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = hex_fill("2C3E50")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 28

    stages = [
        ("Stage", "Step", "Input Count", "Output Count", "Logic", "Assumption"),
        ("1 - Ingestion", "Load raw JSON", 16, 16, "Parse structured deal JSON; each row = one article/mention", "All sources treated as unverified until scored"),
        ("2 - Dedup", "Exact group match", 16, 14, "Group by duplicate_group field; keep is_primary=True record", "Duplicate group assigned on acquirer+target identity"),
        ("2 - Dedup", "Headline similarity", 14, 14, "SequenceMatcher ratio ≥ 0.75 triggers secondary dedup", "Lower threshold = more aggressive removal"),
        ("3 - Scoring", "FMCG relevance", 14, 14, "Keyword density (30+ FMCG terms) × category fit × deal type", "Score ≥5.5 composite required for newsletter inclusion"),
        ("3 - Scoring", "Credibility tier", 14, 14, "T1=SEC/Official (9-10), T2=Trade/Advisory (6-7), T3=Blog (2-4)", "T1 sources given 40% weight in composite score"),
        ("4 - Newsletter", "Categorise & draft", 14, "5 sections", "Deals split: Mega >$10B / Strategic $1-10B / Bolt-on / PE / Failed", "Values at announcement; undisclosed = no public figure"),
    ]
    for r, row_data in enumerate(stages, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 2:
                cell.font = Font(bold=True, size=10, color="FFFFFF")
                cell.fill = hex_fill("34495E")
            else:
                cell.font = Font(size=9)
                cell.fill = hex_fill("ECF0F1" if r % 2 == 0 else "FFFFFF")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = thin_border()
        ws2.row_dimensions[r].height = 40
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 50
    ws2.column_dimensions["F"].width = 40

    # ── Sheet 3: Raw Data ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Raw Data")
    raw_headers = ["id","headline","acquirer","target","deal_value_usd_bn","deal_type",
                   "category","announced_date","status","geography","source",
                   "source_credibility","relevance_score","credibility_score",
                   "composite_score","include_in_newsletter","strategic_rationale"]
    for j, h in enumerate(raw_headers, 1):
        c = ws3.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = hex_fill("2C3E50")
        c.border = thin_border()
    with open(deals_path, encoding="utf-8") as f:
        all_deals = json.load(f)
    for i, deal in enumerate(all_deals, 2):
        for j, key in enumerate(raw_headers, 1):
            v = deal.get(key, "")
            ws3.cell(row=i, column=j, value=str(v) if isinstance(v, list) else v).border = thin_border()
    for j in range(1, len(raw_headers)+1):
        ws3.column_dimensions[get_column_letter(j)].width = 18
    ws3.column_dimensions["B"].width = 42
    ws3.column_dimensions["Q"].width = 50

    # ── Sheet 4: Chart ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Deal Value Chart")
    ws4["A1"] = "Deal"
    ws4["B1"] = "Value ($B)"
    chart_deals = [(d["headline"][:35], d.get("deal_value_usd_bn", 0)) for d in included if d.get("deal_value_usd_bn")]
    chart_deals.sort(key=lambda x: x[1], reverse=True)
    for i, (name, val) in enumerate(chart_deals, 2):
        ws4.cell(row=i, column=1, value=name)
        ws4.cell(row=i, column=2, value=val)

    chart = BarChart()
    chart.type = "bar"
    chart.title = "FMCG M&A Deal Values 2025–2026 ($B)"
    chart.y_axis.title = "Deal Value ($B)"
    chart.style = 10
    data = Reference(ws4, min_col=2, min_row=1, max_row=len(chart_deals)+1)
    cats = Reference(ws4, min_col=1, min_row=2, max_row=len(chart_deals)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 22
    chart.height = 14
    ws4.add_chart(chart, "D2")

    wb.save(output_path)
    print(f"[Excel] Saved → {output_path}")


if __name__ == "__main__":
    create_excel(
        "deals_final.json",
        "FMCG_MA_Newsletter.xlsx"
    )
